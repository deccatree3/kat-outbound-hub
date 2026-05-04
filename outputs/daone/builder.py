"""
다원 발주서 빌더 — EZA 확장주문검색.xls(신양식, 22컬럼) → 다원 표준 발주서.xlsx.

신양식 EZA는 다원 발주서와 거의 동일한 컬럼 구성을 가짐. 변환 규칙:

  EZA 컬럼              →  다원 컬럼
  --------------------     -----------------
  판매처그룹              →  (drop, 제품코드 분기 조건만 사용)
  몰명(또는 몰코드)        →  몰명(또는 몰코드)  (빈값 → "000000000001")
  출하의뢰번호            →  출하의뢰번호
  출하의뢰항번            →  출하의뢰항번
  주문번호               →  주문번호
  상품명                 →  상품명
  제품코드               →  제품코드  (빈값 → 판매처그룹="캐처스"이면 상품메모, 그 외는 바코드)
  바코드                 →  (제품코드 fallback)
  상품메모               →  (캐처스 제품코드 fallback)
  상품수량               →  주문수량
  주문자이름             →  주문자명
  주문자연락처1           →  주문자연락처1
  주문자연락처2           →  주문자연락처2
  수취인명               →  수취인명
  수취인연락처1           →  수취인연락처1
  수취인연락처2           →  수취인연락처2
  수취인우편번호          →  수취인우편번호
  수취인주소1             →  수취인주소1
  주소2                  →  주소2  (빈값 → 수취인주소1 복사)
  배송메시지              →  배송메시지
  송장번호                →  송장번호
  택배사명                →  택배사명

출력은 단일 `발주서` 시트, 다원 19컬럼.
"""
import io
from typing import Dict, List

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import xlrd

# 다원 발주서 헤더 행 배경색 (#E8E8E8)
_HEADER_FILL = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
# KSE 큐텐 국내 빌드 시 'NO' 추가 컬럼 헤더 색상 (#FFFF00)
_NO_COL_HEADER_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# 인박스NO 단위 격번 색상 — 단일 색 (예시 파일 기준).
# 같은 인박스(=같은 사람 주문=같은 KSE 송장) 모든 행 같은 색.
# 패턴: 인박스NO 1=색없음, 2=색, 3=색없음, 4=색 ... (홀수 색없음, 짝수 색)
# 색 = Office Accent6, Lighter 80% (#FCE4D6 연한 살구)
_GROUP_FILL_COLOR = 'FCE4D6'


DAONE_HEADERS = [
    '몰명(또는 몰코드)',
    '출하의뢰번호',
    '출하의뢰항번',
    '주문번호',
    '상품명',
    '제품코드',
    '주문수량',
    '주문자명',
    '주문자연락처1',
    '주문자연락처2',
    '수취인명',
    '수취인연락처1',
    '수취인연락처2',
    '수취인우편번호',
    '수취인주소1',
    '주소2',
    '배송메시지',
    '송장번호',
    '택배사명',
]

# 다원 몰코드 기본값 (EZA가 비워서 보냈을 때 채움)
DEFAULT_몰코드 = '000000000001'

# 헤더 직접 매핑 (분기 규칙은 transform_to_daone에서 별도 처리)
EZA_TO_DAONE = {
    '몰명(또는 몰코드)':  '몰명(또는 몰코드)',
    '출하의뢰번호':       '출하의뢰번호',
    '출하의뢰항번':       '출하의뢰항번',
    '주문번호':           '주문번호',
    '상품명':             '상품명',
    '제품코드':           '제품코드',
    '상품수량':           '주문수량',
    '주문자이름':         '주문자명',
    '주문자연락처1':      '주문자연락처1',
    '주문자연락처2':      '주문자연락처2',
    '수취인명':           '수취인명',
    '수취인연락처1':      '수취인연락처1',
    '수취인연락처2':      '수취인연락처2',
    '수취인우편번호':     '수취인우편번호',
    '수취인주소1':        '수취인주소1',
    '주소2':              '주소2',
    '배송메시지':         '배송메시지',
    '송장번호':           '송장번호',
    '택배사명':           '택배사명',
}


def _cell_str(value, ctype) -> str:
    """xlrd 셀 값 → 문자열. NUMBER 정수형 float은 정수 표기.
    텍스트 컬럼(우편번호/전화 등)의 leading-zero는 EZA가 TEXT로 보내므로 보존됨.
    """
    if value is None or value == '':
        return ''
    if ctype == 2:  # NUMBER
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)
    return str(value)


def parse_eza_xls(data: bytes) -> List[Dict]:
    """EZA 확장주문검색.xls bytes → 헤더 기반 dict 리스트.
    첫 시트의 row 0=헤더, row 1+=데이터로 가정.
    """
    wb = xlrd.open_workbook(file_contents=data)
    ws = wb.sheet_by_index(0)
    if ws.nrows < 1:
        return []
    headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
    rows = []
    for r in range(1, ws.nrows):
        d = {h: _cell_str(ws.cell_value(r, c), ws.cell_type(r, c))
             for c, h in enumerate(headers)}
        rows.append(d)
    return rows


def transform_to_daone(eza_rows: List[Dict]) -> List[Dict]:
    """신양식 EZA dict → 다원 dict. 빈값 fallback / 판매처그룹 분기 적용."""
    out = []
    for eza in eza_rows:
        daone = {h: '' for h in DAONE_HEADERS}
        for eza_h, daone_h in EZA_TO_DAONE.items():
            daone[daone_h] = eza.get(eza_h, '')

        # 1) 몰명(또는 몰코드) 빈값 → 기본값
        if not str(daone.get('몰명(또는 몰코드)') or '').strip():
            daone['몰명(또는 몰코드)'] = DEFAULT_몰코드

        # 2) 주소2 빈값 → 수취인주소1 복사
        if not str(daone.get('주소2') or '').strip():
            daone['주소2'] = daone.get('수취인주소1', '')

        # 3) 제품코드 빈값 → 판매처그룹 분기
        if not str(daone.get('제품코드') or '').strip():
            group = str(eza.get('판매처그룹') or '').strip()
            if group == '캐처스':
                daone['제품코드'] = eza.get('상품메모', '')
            else:
                daone['제품코드'] = eza.get('바코드', '')

        # 주문수량 정수 보정
        q = daone.get('주문수량')
        if q not in (None, '', 0):
            try:
                daone['주문수량'] = int(float(q))
            except (ValueError, TypeError):
                pass

        out.append(daone)
    return out


def build_daone_xlsx(daone_rows: List[Dict],
                     add_packing_columns: bool = False) -> bytes:
    """다원 발주서.xlsx bytes 생성. 단일 `발주서` 시트, 19 컬럼.

    add_packing_columns=True 면 마지막에 4 컬럼 추가 (모두 헤더 #FFFF00):
      인박스 / 인박스NO / 아웃박스 / 아웃박스NO

    각 행이 '_group_key' 메타 키(예: (도착지송장번호, 장바구니번호))를 가지고 있어야
    인박스NO가 그룹별로 부여됨. 패킹 계산은 outputs.packing.boxes:compute_packing.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '발주서'
    extra = ['인박스', '인박스NO', '아웃박스', '아웃박스NO'] if add_packing_columns else []
    headers = list(DAONE_HEADERS) + extra
    ws.append(headers)
    # 표준 헤더 색상
    for c in range(1, len(DAONE_HEADERS) + 1):
        ws.cell(1, c).fill = _HEADER_FILL
    # 패킹 컬럼 헤더 색상 (#FFFF00)
    if add_packing_columns:
        for c in range(len(DAONE_HEADERS) + 1, len(headers) + 1):
            ws.cell(1, c).fill = _NO_COL_HEADER_FILL

    # ─── 패킹 계산 (compute_packing 으로 위임) ───
    if add_packing_columns:
        from outputs.packing.boxes import compute_packing
        ordered = compute_packing(daone_rows)
    else:
        ordered = daone_rows

    group_fill = PatternFill(start_color=_GROUP_FILL_COLOR,
                             end_color=_GROUP_FILL_COLOR, fill_type='solid')

    for ri, r in enumerate(ordered, 2):
        row_values = [r.get(h, '') for h in DAONE_HEADERS]
        if add_packing_columns:
            row_values += [r.get('_packing_inbox'), r.get('_packing_inbox_no'),
                           r.get('_packing_outbox'), r.get('_packing_outbox_no')]
        ws.append(row_values)
        # 인박스NO 짝수면 행 전체에 색 채움 (같은 인박스의 모든 SKU 행 같은 색)
        if add_packing_columns:
            inbno = r.get('_packing_inbox_no')
            if isinstance(inbno, int) and inbno % 2 == 0:
                for c in range(1, len(headers) + 1):
                    ws.cell(ri, c).fill = group_fill

    widths = [14, 18, 14, 14, 40, 14, 8, 12, 16, 16, 12, 16, 16, 12, 50, 50, 30, 16, 12]
    if add_packing_columns:
        widths = widths + [22, 8, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def convert_eza_to_daone(eza_bytes: bytes) -> tuple[bytes, int]:
    """원샷 변환: EZA xls bytes → 다원 xlsx bytes. (xlsx_bytes, row_count) 반환."""
    eza_rows = parse_eza_xls(eza_bytes)
    daone_rows = transform_to_daone(eza_rows)
    return build_daone_xlsx(daone_rows), len(daone_rows)


# ─── KSE OMS 주문내역 (큐텐 국내출고, 한국 KSE 경유) ───────────────────────
# KSE OMS 다운로드 양식 (26컬럼):
#   번호 / 등록일 / 접수번호 / 배송상태 / 배송타입 / 도착지송장번호 / 판매마켓 /
#   주문일 / 주문번호 / 장바구니번호 / 상품코드 / 판매자코드 /
#   상품명(판매마켓대표상품명) / 옵션명 / 옵션코드 / 금액 / 수량 /
#   받는사람 / 받는사람전화 / 우편번호 / 주소 / 사이즈 / 실무게 / 부피무게 / 적용무게 / RegionName

def parse_kse_oms_xlsx(data: bytes) -> List[Dict]:
    """KSE OMS 주문내역.xlsx bytes → 헤더 기반 dict 리스트.
    첫 시트의 row 1=헤더, row 2+=데이터.
    """
    import io as _io
    import openpyxl as _opx
    wb = _opx.load_workbook(_io.BytesIO(data), data_only=True)
    ws = wb.active
    if ws.max_row < 2:
        return []
    headers = [str(ws.cell(1, c).value).strip() if ws.cell(1, c).value is not None else ''
               for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        d = {}
        for c, h in enumerate(headers, 1):
            v = ws.cell(r, c).value
            if v is None:
                d[h] = ''
            elif isinstance(v, float) and v.is_integer():
                d[h] = str(int(v))
            else:
                d[h] = str(v)
        # 빈 행 건너뜀 (번호 컬럼이 비어있으면)
        if not d.get('번호') and not d.get('주문번호'):
            continue
        rows.append(d)
    return rows


def kse_oms_to_daone(kse_rows: List[Dict]) -> List[Dict]:
    """KSE OMS dict 리스트 → 다원 19컬럼 dict 리스트.
    SKU 매핑(제품코드)은 추후 단계에서 추가 — 현재는 빈값.
    """
    out = []
    for k in kse_rows:
        receiver = k.get('받는사람', '')
        phone = k.get('받는사람전화', '')
        addr = k.get('주소', '')
        zip_code = k.get('우편번호', '')
        name = k.get('상품명(판매마켓대표상품명)', '')
        option = k.get('옵션명', '')
        full_name = name + (' / ' + option if option else '')

        # 수량 정수
        try:
            qty = int(float(k.get('수량', 0))) if k.get('수량') else 0
        except (ValueError, TypeError):
            qty = 0

        d = {h: '' for h in DAONE_HEADERS}
        d['몰명(또는 몰코드)'] = DEFAULT_몰코드
        d['출하의뢰번호']     = k.get('판매마켓', '')
        d['출하의뢰항번']     = k.get('주문번호', '')      # 큐텐 주문번호
        d['주문번호']         = k.get('접수번호', '')      # KSE 송장(접수)번호
        d['상품명']           = full_name.strip()
        d['제품코드']         = ''  # SKU 매핑은 다음 단계
        d['주문수량']         = qty
        d['주문자명']         = receiver
        d['주문자연락처1']    = phone
        d['주문자연락처2']    = ''
        d['수취인명']         = receiver
        d['수취인연락처1']    = phone
        d['수취인연락처2']    = ''
        d['수취인우편번호']   = zip_code
        d['수취인주소1']      = addr
        d['주소2']            = addr  # 빈값 fallback 규칙 (수취인주소1 복사)
        d['배송메시지']       = ''
        d['송장번호']         = k.get('도착지송장번호', '')
        d['택배사명']         = k.get('배송타입', '')
        out.append(d)
    return out


def convert_kse_oms_to_daone(xlsx_bytes: bytes) -> tuple[bytes, int]:
    """원샷 변환 (매핑 미사용, 제품코드 빈값): KSE OMS xlsx → 다원 19컬럼 발주서.
    호환성용. 실제 운영에선 kse_oms_to_daone_with_mapping 사용 권장.
    """
    kse_rows = parse_kse_oms_xlsx(xlsx_bytes)
    daone_rows = kse_oms_to_daone(kse_rows)
    return build_daone_xlsx(daone_rows), len(daone_rows)


# KSE 한국 집하지 (다원이 큐텐 국내 출고 시 보내는 곳)
# 인박스에 KSE 송장(PDF) 부착 → 아웃박스로 합포장 → 이 집하지로 발송 → KSE 가 한국→일본 이동
KSE_KR_DEPOT = {
    'name':       'KSE',
    'phone':      '02 3143 5555',
    'zip':        '03917',
    'address':    '서울특별시 마포구 구룡길 36, (주)국제로지스틱 수색 EC 물류센터 내 G1 (GATE 22)',
    'msg':        'KSE',
}


def kse_oms_to_daone_with_mapping(kse_rows: List[Dict], mappings: Dict) -> Dict:
    """KSE OMS dict + 채널 매핑 (channel='cachers_qoo10_kr') → 다원 19컬럼.

    분기:
      매핑 없음                   → unknown_rows (등록 강제)
      매핑 + sku_codes 정상       → daone_rows 에 1→N 펼침
      매핑 + sku_codes='-'/빈     → incomplete_rows (다원 SKU 미입력)

    반환: dict {
        'daone_rows': [...], 'unknown_rows': [...], 'incomplete_rows': [...],
    }
    """
    daone_rows: List[Dict] = []
    unknown_rows: List[Dict] = []
    incomplete_rows: List[Dict] = []

    for k in kse_rows:
        name = (k.get('상품명(판매마켓대표상품명)') or '').strip()
        option = (k.get('옵션명') or '').strip()
        m = mappings.get((name, option))

        info = {
            '주문번호': k.get('주문번호', ''),
            '접수번호': k.get('접수번호', ''),
            '상품명': name,
            '옵션명': option,
            '수량': k.get('수량', ''),
        }

        if m is None:
            unknown_rows.append(info)
            continue
        valid = [(s.strip(), q) for s, q in zip(m.get('sku_codes', []), m.get('quantities', []))
                 if s and s.strip() and s.strip() != '-']
        if not valid:
            incomplete_rows.append(info)
            continue

        try:
            base_qty = int(float(k.get('수량', 1) or 1))
        except (ValueError, TypeError):
            base_qty = 1

        full_name = name + (' / ' + option if option else '')

        for sku_code, sku_unit in valid:
            try:
                unit = int(sku_unit)
            except (ValueError, TypeError):
                unit = 1
            d = {h: '' for h in DAONE_HEADERS}
            d['몰명(또는 몰코드)'] = DEFAULT_몰코드
            d['출하의뢰번호']     = k.get('판매마켓', '')
            d['출하의뢰항번']     = k.get('주문번호', '')      # 큐텐 주문번호
            d['주문번호']         = k.get('접수번호', '')      # KSE 송장(접수)번호
            d['상품명']           = full_name.strip()
            d['제품코드']         = sku_code
            d['주문수량']         = unit * base_qty
            # 패킹 그룹 키 보존 (도착지송장번호 + 장바구니번호 → 같은 인박스NO)
            d['_group_key']      = (str(k.get('도착지송장번호', '')),
                                    str(k.get('장바구니번호', '')))
            # 다원 → KSE 한국 집하지 고정 정보 (일본 고객 정보 아님)
            # 도착지송장번호 + KSE 송장 PDF 는 인박스에 부착되어 KSE 가 일본으로 이동.
            d['주문자명']         = KSE_KR_DEPOT['name']
            d['주문자연락처1']    = KSE_KR_DEPOT['phone']
            d['주문자연락처2']    = KSE_KR_DEPOT['phone']
            d['수취인명']         = KSE_KR_DEPOT['name']
            d['수취인연락처1']    = KSE_KR_DEPOT['phone']
            d['수취인연락처2']    = KSE_KR_DEPOT['phone']
            d['수취인우편번호']   = KSE_KR_DEPOT['zip']
            d['수취인주소1']      = KSE_KR_DEPOT['address']
            d['주소2']            = KSE_KR_DEPOT['address']
            d['배송메시지']       = KSE_KR_DEPOT['msg']
            # 송장번호/택배사명 은 다원이 채움 (한국 내 다원→KSE 집하지 운송 송장)
            d['송장번호']         = ''
            d['택배사명']         = ''
            daone_rows.append(d)

    return {
        'daone_rows': daone_rows,
        'unknown_rows': unknown_rows,
        'incomplete_rows': incomplete_rows,
    }


# ─── 캐처스 메이커스 (카카오메이커스) ─────────────────────────────────────

# 다원 메이커스 발주서 샘플 기준 fixed 값
MAKERS_DAONE_출하의뢰번호 = '[캐처스] 카카오메이커스'

# 메이커스 주문서 헤더 (시트명 '주문내역', 22컬럼)
MAKERS_HEADERS = [
    '배송번호', '결제번호', '주문번호', '회차상품 번호', '상품', '옵션', '수량',
    '주문금액', '배송비', '택배사명', '택배사코드', '송장번호(하이픈 없이 입력)',
    '주문일시', '결제일시', '수령인명', '수령인 연락처1', '수령인 연락처2',
    '배송주소', '배송메시지', '우편번호', '정산방식', '발주상태',
]


def parse_makers_xlsx(xlsx_bytes: bytes) -> List[Dict]:
    """카카오메이커스 주문내역.xlsx → list of dict.
    시트명 '주문내역' 사용. 헤더 첫 행, 데이터 둘째 행부터.
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    sheet_name = '주문내역' if '주문내역' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    header = [str(h).strip() if h is not None else '' for h in header_row]

    out: List[Dict] = []
    for row in rows_iter:
        if row is None or all(c is None or str(c).strip() == '' for c in row):
            continue
        d = {}
        for i, h in enumerate(header):
            v = row[i] if i < len(row) else None
            d[h] = '' if v is None else v
        out.append(d)
    return out


def _makers_int(v) -> int:
    try:
        return int(float(v)) if v not in (None, '') else 0
    except (ValueError, TypeError):
        return 0


def _makers_str(v) -> str:
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def makers_to_daone_with_mapping(makers_rows: List[Dict], mappings: Dict) -> Dict:
    """메이커스 주문내역 dict + 채널 매핑 (channel='cachers_makers') → 다원 19컬럼.

    분기:
      매핑 없음               → unknown_rows (등록 강제)
      매핑 + sku_codes 정상   → daone_rows 에 1→N 펼침
      매핑 + sku_codes='-'/빈 → incomplete_rows

    반환: dict {'daone_rows', 'unknown_rows', 'incomplete_rows'}
    """
    daone_rows: List[Dict] = []
    unknown_rows: List[Dict] = []
    incomplete_rows: List[Dict] = []

    for r in makers_rows:
        product = (r.get('상품') or '').strip()
        option = (r.get('옵션') or '').strip()
        m = mappings.get((product, option))

        info = {
            '주문번호':   _makers_str(r.get('주문번호')),
            '배송번호':   _makers_str(r.get('배송번호')),
            '상품':       product,
            '옵션':       option,
            '수량':       _makers_int(r.get('수량')),
            '수령인명':   r.get('수령인명') or '',
        }

        if m is None:
            unknown_rows.append(info)
            continue
        valid = [(s.strip(), q) for s, q in zip(m.get('sku_codes', []), m.get('quantities', []))
                 if s and s.strip() and s.strip() != '-']
        if not valid:
            incomplete_rows.append(info)
            continue

        base_qty = _makers_int(r.get('수량')) or 1
        full_name = product + (' / ' + option if option else '')
        recipient = (r.get('수령인명') or '').strip()
        phone1 = _makers_str(r.get('수령인 연락처1'))
        phone2 = _makers_str(r.get('수령인 연락처2'))
        zip_code = _makers_str(r.get('우편번호'))
        address = (r.get('배송주소') or '').strip()
        msg = (r.get('배송메시지') or '').strip()
        waybill = _makers_str(r.get('송장번호(하이픈 없이 입력)'))
        carrier = (r.get('택배사명') or '').strip()
        order_no = _makers_str(r.get('주문번호'))
        ship_no = _makers_str(r.get('배송번호'))

        for sku_code, sku_unit in valid:
            try:
                unit = int(sku_unit)
            except (ValueError, TypeError):
                unit = 1
            d = {h: '' for h in DAONE_HEADERS}
            d['몰명(또는 몰코드)'] = DEFAULT_몰코드
            d['출하의뢰번호']     = MAKERS_DAONE_출하의뢰번호
            d['출하의뢰항번']     = ship_no
            d['주문번호']     = order_no
            d['상품명']           = full_name.strip()
            d['제품코드']         = sku_code
            d['주문수량']         = unit * base_qty
            d['주문자명']         = recipient
            d['주문자연락처1']    = phone1
            d['주문자연락처2']    = phone2
            d['수취인명']         = recipient
            d['수취인연락처1']    = phone1
            d['수취인연락처2']    = phone2
            d['수취인우편번호']   = zip_code
            d['수취인주소1']      = address
            d['주소2']            = ''
            d['배송메시지']       = msg
            d['송장번호']         = waybill
            d['택배사명']         = carrier
            daone_rows.append(d)

    return {
        'daone_rows': daone_rows,
        'unknown_rows': unknown_rows,
        'incomplete_rows': incomplete_rows,
    }

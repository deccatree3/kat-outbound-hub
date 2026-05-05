"""
캐처스 3PL 출고요청서 빌더.

EZA 확장주문검색.xls (신양식 31컬럼) → 캐처스-3PL-참기름-자연앤미 출고요청서.xlsx (25컬럼).

필터: 업로드 파일의 '공급처' 컬럼이 정확히 TARGET_SUPPLIER 인 행만 추출.
"""
import io
from typing import Dict, List, Tuple

import openpyxl
from openpyxl.utils import get_column_letter


TARGET_SUPPLIER = '캐처스-3PL-참기름-자연앤미'

OUTPUT_HEADERS = [
    '공급처', '주문일', '주문시간', '발주일', '발주시간', '몰명',
    '출하의뢰번호', '출하의뢰항번', '주문번호',
    '판매처 상품명', '판매처 옵션', '자체상품코드', '주문수량',
    '주문자이름', '주문자연락처2', '주문자연락처1',
    '수취인명', '수취인연락처2', '수취인연락처1',
    '수취인우편번호', '수취인주소1',
    '배송메시지', 'CS', '송장번호', '택배사',
]

# 출고요청서 컬럼 → EZA 컬럼 매핑.
# 키 = 출고요청서 컬럼 (왼→오 순). 값 = EZA 컬럼명. None 이면 빈값.
OUTPUT_TO_EZA: Dict[str, str] = {
    '공급처':         '공급처',
    '주문일':         '주문일',
    '주문시간':       '주문시간',
    '발주일':         '발주일',
    '발주시간':       '발주시간',
    '몰명':           None,         # EZA 에 없음 → 빈값
    '출하의뢰번호':    '출하의뢰번호',
    '출하의뢰항번':    '출하의뢰항번',
    '주문번호':       '주문번호',
    '판매처 상품명':   '판매처 상품명',
    '판매처 옵션':    '판매처 옵션',
    '자체상품코드':    '제품코드',
    '주문수량':       '주문수량',
    '주문자이름':     '주문자이름',
    '주문자연락처2':   '주문자연락처2',
    '주문자연락처1':   '주문자연락처1',
    '수취인명':       '수취인명',
    '수취인연락처2':   '수취인연락처2',
    '수취인연락처1':   '수취인연락처1',
    '수취인우편번호':  '수취인우편번호',
    '수취인주소1':    '수취인주소1',
    '배송메시지':     '배송메시지',
    'CS':             'CS',
    '송장번호':       '송장번호',
    '택배사':         '택배사명',
}


def filter_target_rows(eza_rows: List[Dict]) -> List[Dict]:
    """공급처 == TARGET_SUPPLIER 인 행만 반환."""
    return [r for r in eza_rows
            if str(r.get('공급처', '')).strip() == TARGET_SUPPLIER]


def build_cachers_3pl_xlsx(eza_rows: List[Dict]) -> Tuple[bytes, int]:
    """필터된 EZA dict 들을 25컬럼 출고요청서.xlsx 로 빌드.
    반환: (xlsx_bytes, target_row_count).
    """
    target_rows = filter_target_rows(eza_rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws.append(OUTPUT_HEADERS)

    for r in target_rows:
        out_row = []
        for h in OUTPUT_HEADERS:
            eza_key = OUTPUT_TO_EZA.get(h)
            v = r.get(eza_key, '') if eza_key else ''
            out_row.append(v)
        ws.append(out_row)

    # 컬럼 폭 (운영 시 가독성)
    widths = [22, 12, 10, 12, 10, 10, 22, 18, 14,
              30, 30, 18, 8, 14, 16, 16, 14, 16, 16, 12, 50, 30, 8, 16, 12]
    for i, w in enumerate(widths, 1):
        if i <= len(OUTPUT_HEADERS):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), len(target_rows)

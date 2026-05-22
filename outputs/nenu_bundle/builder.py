"""
네뉴 일반주문 번들작업파일 빌더 — EZA 확장주문검색.xls → 일반주문 번들작업건.xlsx.

마스터 양식 (`outputs/nenu_bundle/template.xlsx`):
  - 시트 1개: `form` (단품 + 세트 마스터 표 + 수식)
  - 7컬럼: 바코드 / 상품명 / 출고수량 / 입고수량 / 세트인치 / =D*E / 모체단품명
  - 단품 행: C에 SUMIFS 수식 (같은 시트 G,F 참조), D='#' (입력 금지)
  - 세트 행: C='#', D는 빈값 (사용자가 채울 칸), E=인치, F=`=D*E`, G=모체단품명

빌드 흐름:
  1. EZA 확장주문검색.xls 에서 바코드별 상품수량 합계 산출 (parse_eza_for_bundle)
  2. 마스터 template.xlsx 로드, `form` 시트 복사해서 새 시트 추가
  3. 새 시트의 세트 행 D셀에 EZA 합계 정수 입력 (SUMIFS 외부참조 대체)
  4. 단품 C셀의 SUMIFS는 그대로 두면 Excel이 같은 시트 F/G로 자동 계산
  5. xlsx bytes 반환

참고: 자매 프로젝트 nn-rocketgrowth_inventory와 차이 — 그쪽은 사용자가 매월 작업 파일을 줌. 이 프로젝트는 마스터 양식만 보관하고 매일 EZA만 받음.
"""
import io
import os
import datetime
from typing import Dict

import openpyxl
import xlrd

_THIS = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(_THIS, 'template.xlsx')


def _normalize_barcode(value) -> str:
    """xlrd/openpyxl이 NUMBER로 읽은 바코드 → 정수 표기 string."""
    if value is None or value == '':
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    s = str(value).strip()
    if s.endswith('.0'):
        s = s[:-2]
    return s


GIFT_KEYWORD = '선물세트'


def load_master_parent_names() -> list:
    """모체 단품명 후보 — 템플릿 form 시트 **G열(2행~) 고유값** + DB 오버레이 parent. 정렬.

    추가폼의 '모체 단품명' 드롭다운(검색·선택)용. 1행(헤더) 제외, 중복 제거.
    """
    names = set()
    try:
        if os.path.exists(TEMPLATE_PATH):
            wb = openpyxl.load_workbook(TEMPLATE_PATH, read_only=True)
            if 'form' in wb.sheetnames:
                ws = wb['form']
                for row in ws.iter_rows(min_row=2, min_col=7, max_col=7, values_only=True):
                    v = row[0]
                    if v is not None and str(v).strip():
                        names.add(str(v).strip())
            wb.close()
    except Exception:
        pass
    try:
        from db import nenu_bundle_extra as _nbe
        for ex in _nbe.load_all():
            p = (ex.get('parent_name') or '').strip()
            if p:
                names.add(p)
    except Exception:
        pass
    return sorted(names)


def parse_eza_for_bundle(data: bytes, exclude_groups=('캐처스',)) -> tuple[Dict[str, int], Dict[str, str]]:
    """이지어드민 확장주문검색.xls bytes → ({바코드: 상품수량 합계}, {바코드: 상품명}).

    필터 (둘 다 적용):
      - 판매처그룹 ∈ exclude_groups (기본 '캐처스') 제외 — 번들은 네뉴 전용
      - 상품명에 '선물세트' 미포함 행 제외 — 번들작업은 선물세트 건만 해당
    """
    wb = xlrd.open_workbook(file_contents=data)
    ws = wb.sheet_by_index(0)
    if ws.nrows < 1:
        return {}
    headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
    try:
        bar_idx = headers.index('바코드')
        qty_idx = headers.index('상품수량')
    except ValueError as e:
        raise RuntimeError(f"이지어드민 헤더에서 '바코드' 또는 '상품수량' 컬럼을 찾지 못했습니다 ({e})")
    try:
        name_idx = headers.index('상품명')
    except ValueError as e:
        raise RuntimeError(f"이지어드민 헤더에서 '상품명' 컬럼을 찾지 못했습니다 ({e})")
    grp_idx = headers.index('판매처그룹') if '판매처그룹' in headers else None
    excluded = set(exclude_groups or ())

    totals: Dict[str, int] = {}
    names: Dict[str, str] = {}
    for r in range(1, ws.nrows):
        if grp_idx is not None:
            g = str(ws.cell_value(r, grp_idx)).strip()
            if g in excluded:
                continue
        name = str(ws.cell_value(r, name_idx) or '')
        if GIFT_KEYWORD not in name:
            continue
        bar = _normalize_barcode(ws.cell_value(r, bar_idx))
        qty_raw = ws.cell_value(r, qty_idx)
        try:
            qty = int(float(qty_raw)) if qty_raw not in (None, '') else 0
        except (ValueError, TypeError):
            qty = 0
        if bar:
            totals[bar] = totals.get(bar, 0) + qty
            names.setdefault(bar, name.strip())
    return totals, names


def build_bundle_xlsx(eza_bytes,
                      work_date: datetime.date,
                      sequence: int) -> tuple[bytes, Dict]:
    """마스터 양식에 EZA 합계 채워서 일반주문 번들작업건.xlsx bytes 반환.
    eza_bytes: bytes 또는 List[bytes] (여러 파일 합산).
    """
    if not os.path.exists(TEMPLATE_PATH):
        raise RuntimeError(f"마스터 템플릿이 없습니다: {TEMPLATE_PATH}")

    eza_totals: Dict[str, int] = {}
    eza_names: Dict[str, str] = {}
    sources = eza_bytes if isinstance(eza_bytes, (list, tuple)) else [eza_bytes]
    for b in sources:
        t, n = parse_eza_for_bundle(b)
        for bar, qty in t.items():
            eza_totals[bar] = eza_totals.get(bar, 0) + qty
        for bar, nm in n.items():
            eza_names.setdefault(bar, nm)

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    if 'form' not in wb.sheetnames:
        raise RuntimeError("템플릿에 'form' 시트가 없습니다")

    # 외부 워크북 참조 ([1]Worksheet) 메타 제거 — 우리가 D셀을 정수로 덮어써서
    # 더 이상 참조되지 않으면 Excel이 corrupt(외부 수식 참조에서 캐시된 값) 경고를 띄움.
    if hasattr(wb, '_external_links'):
        wb._external_links = []

    new_sheet_name = f"{work_date.strftime('%y%m%d')}_{sequence}차"
    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]
    ws = wb.copy_worksheet(wb['form'])
    ws.title = new_sheet_name
    # 새 시트를 첫 번째로 이동
    wb.move_sheet(new_sheet_name, offset=-len(wb.sheetnames) + 1)
    # form 시트 숨김 (마스터 백업)
    wb['form'].sheet_state = 'hidden'

    # 마스터 행 분류:
    #   세트 행 = col C='#' 이면서 col E(세트인치) 가 숫자
    #   단품 행 = col C 가 SUMIFS 수식이면서 col D='#'
    # 모체 단품명 → 단품 행 매핑도 같이 기록 (visible 행 결정용)
    set_barcodes = set()
    single_barcodes = set()
    set_row_index_by_bar = {}
    parent_name_by_set_row = {}
    single_row_by_name = {}
    for row in range(2, ws.max_row + 1):
        a = ws.cell(row, 1).value
        b_val = ws.cell(row, 2).value
        c_val = ws.cell(row, 3).value
        d_val = ws.cell(row, 4).value
        e_val = ws.cell(row, 5).value
        g_val = ws.cell(row, 7).value
        bar = _normalize_barcode(a)
        if not bar:
            continue
        if str(c_val).strip() == '#' and isinstance(e_val, (int, float)):
            set_barcodes.add(bar)
            set_row_index_by_bar[bar] = row
            if g_val:
                parent_name_by_set_row[row] = str(g_val).strip()
        elif str(d_val).strip() == '#':
            single_barcodes.add(bar)
            if b_val:
                single_row_by_name[str(b_val).strip()] = row

    # ── DB 오버레이 병합 — 마스터 템플릿에 없는 신규 선물세트 세트 행을 append ──
    # (template.xlsx 는 git 레포 파일이라 런타임 직접수정 불가 → DB 에 보관, 여기서 합성)
    try:
        from db import nenu_bundle_extra as _nbe
        _extras = _nbe.load_all()
    except Exception:
        _extras = []
    _append_row = ws.max_row + 1
    _overlay_parents = set()
    for ex in _extras:
        bar = _normalize_barcode(ex.get('barcode'))
        if not bar or bar in set_barcodes:
            continue  # 템플릿/이미 추가분에 있으면 skip
        ws.cell(_append_row, 1, ex.get('barcode'))
        ws.cell(_append_row, 2, ex.get('product_name'))
        ws.cell(_append_row, 3, '#')
        ws.cell(_append_row, 5, int(ex.get('set_units') or 1))
        ws.cell(_append_row, 6, f"=D{_append_row}*E{_append_row}")
        parent = (ex.get('parent_name') or '').strip()
        if parent:
            ws.cell(_append_row, 7, parent)
            parent_name_by_set_row[_append_row] = parent
            _overlay_parents.add(parent)
        set_barcodes.add(bar)
        set_row_index_by_bar[bar] = _append_row
        _append_row += 1

    # 오버레이 세트의 모체 단품이 템플릿 단품에 없으면 단품 행도 append
    # (단품 출고수량 C=SUMIFS 가 그 모체의 세트들을 자동 집계하도록)
    for parent in sorted(_overlay_parents):
        if parent in single_row_by_name:
            continue
        ws.cell(_append_row, 2, parent)
        ws.cell(_append_row, 3, f"=SUMIFS($F:$F,$G:$G,B{_append_row})")
        ws.cell(_append_row, 4, '#')
        single_row_by_name[parent] = _append_row
        _append_row += 1

    # 세트 D 채움 + visible 결정
    set_rows_filled = 0
    total_qty = 0
    visible_rows = set()
    parent_names_with_orders = set()
    for bar, row in set_row_index_by_bar.items():
        qty = eza_totals.get(bar, 0)
        ws.cell(row, 4).value = qty
        if qty:
            set_rows_filled += 1
            total_qty += qty
            visible_rows.add(row)
            parent = parent_name_by_set_row.get(row)
            if parent:
                parent_names_with_orders.add(parent)

    # 단품 행: 그 단품을 모체로 하는 세트 중 하나라도 채워졌다면 visible
    for name in parent_names_with_orders:
        row = single_row_by_name.get(name)
        if row:
            visible_rows.add(row)

    # 매칭되지 않은 행 hidden
    for row in range(2, ws.max_row + 1):
        if row not in visible_rows:
            ws.row_dimensions[row].hidden = True

    # auto_filter 메타는 깨끗하게 — 사용자가 Excel에서 직접 사용 가능
    ws.auto_filter.ref = f"A1:G{ws.max_row}"

    eza_bars = set(eza_totals.keys())
    set_matched = sorted(eza_bars & set_barcodes)
    single_matched = sorted(eza_bars & single_barcodes)
    unmatched = sorted(eza_bars - set_barcodes - single_barcodes)
    # 미매칭 상세 (얼럿/템플릿 추가폼용) — 상품명·수량 동반
    unmatched_detail = [
        {'barcode': b, 'name': eza_names.get(b, ''), 'qty': eza_totals.get(b, 0)}
        for b in unmatched
    ]

    buf = io.BytesIO()
    wb.save(buf)
    info = {
        'sheet_name': new_sheet_name,
        'set_rows_filled': set_rows_filled,
        'total_qty': total_qty,
        'set_matched_barcodes': set_matched,
        'single_matched_barcodes': single_matched,
        'unmatched_barcodes': unmatched,
        'unmatched_detail': unmatched_detail,
        'eza_total_rows': sum(1 for v in eza_totals.values() if v),
        'eza_total_qty': sum(eza_totals.values()),
        'master_set_count': len(set_barcodes),
        'master_single_count': len(single_barcodes),
    }
    return buf.getvalue(), info

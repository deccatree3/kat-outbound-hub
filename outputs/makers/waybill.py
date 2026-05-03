"""
메이커스 송장 기입 빌더.

다원 채번.xls (12컬럼) + 메이커스 원본 주문서.xlsx (22컬럼) →
송장번호 채운 메이커스 주문서.xlsx (메이커스 어드민에 그대로 업로드 가능).

매칭 키:
  메이커스 (수령인명, 수령인 연락처1 정규화)
  ↔ 다원 채번 (수취인, 전화번호 또는 핸드폰 정규화)

운송장번호의 하이픈/공백은 제거 (메이커스 양식 컬럼명: '송장번호(하이픈 없이 입력)').
"""
import io
from typing import Dict, List, Tuple

import openpyxl
import xlrd


WAYBILL_COL_HEADER = '송장번호(하이픈 없이 입력)'


def _normalize_phone(value) -> str:
    if value is None:
        return ''
    s = str(value)
    return ''.join(c for c in s if c.isdigit())


def _normalize_waybill(value) -> str:
    if value is None:
        return ''
    s = str(value).strip()
    return ''.join(c for c in s if c.isdigit())


def parse_daone_waybill_xls(data: bytes) -> Dict[Tuple[str, str], List[str]]:
    """다원 채번.xls → {(수취인, 전화정규화): [운송장번호, ...]}.
    같은 키가 여러 행이면 list로 누적.
    """
    wb = xlrd.open_workbook(file_contents=data)
    ws = wb.sheet_by_index(0)
    if ws.nrows < 2:
        return {}
    headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]

    def find_idx(*candidates):
        for c in candidates:
            if c in headers:
                return headers.index(c)
        return None

    name_i = find_idx('수취인', '수령인', '받는사람', '수령인명')
    phone_primary_i = find_idx('전화번호', '연락처1', '전화')
    phone_alt_i = find_idx('핸드폰', '휴대폰', '연락처2')
    waybill_i = find_idx('운송장번호', '송장번호', '송장번호(하이픈 없이 입력)')

    if name_i is None or waybill_i is None:
        raise RuntimeError(
            f"다원 채번 파일에서 '수취인' 또는 '운송장번호' 컬럼을 찾지 못했습니다. "
            f"실제 헤더: {headers}"
        )

    out: Dict[Tuple[str, str], List[str]] = {}
    for r in range(1, ws.nrows):
        name = str(ws.cell_value(r, name_i)).strip()
        phone_p = _normalize_phone(ws.cell_value(r, phone_primary_i)) if phone_primary_i is not None else ''
        phone_a = _normalize_phone(ws.cell_value(r, phone_alt_i)) if phone_alt_i is not None else ''
        # primary 비어있으면 alt 사용
        phone = phone_p or phone_a
        waybill_raw = ws.cell_value(r, waybill_i)
        waybill = _normalize_waybill(waybill_raw)
        if not name or not waybill:
            continue
        key = (name, phone)
        out.setdefault(key, []).append(waybill)
    return out


def fill_makers_waybills(makers_xlsx_bytes: bytes,
                         daone_xls_bytes: bytes) -> Tuple[bytes, Dict]:
    """메이커스 원본.xlsx + 다원 채번.xls → 송장 채워진 메이커스.xlsx + info.

    info: {
        'filled': int,
        'duplicates': List[dict],  # 같은 (수취인, 전화) 키에 채번이 여러 개인 케이스
        'unmatched': List[dict],   # 메이커스 행 중 채번 매칭 실패
        'leftover_waybills': List[Tuple],  # 채번에 있는데 메이커스에 매칭 안 된 키
    }
    """
    waybill_map = parse_daone_waybill_xls(daone_xls_bytes)

    wb = openpyxl.load_workbook(io.BytesIO(makers_xlsx_bytes))
    ws = wb.active  # '주문내역' 가정

    # 헤더 확인 + 컬럼 인덱스
    header_row = [str(c.value).strip() if c.value is not None else '' for c in ws[1]]
    try:
        name_idx = header_row.index('수령인명') + 1
        phone1_idx = header_row.index('수령인 연락처1') + 1
    except ValueError as ex:
        raise RuntimeError(f"메이커스 주문서 헤더에서 컬럼을 찾지 못함: {ex}")
    if WAYBILL_COL_HEADER in header_row:
        waybill_idx = header_row.index(WAYBILL_COL_HEADER) + 1
    elif '송장번호' in header_row:
        waybill_idx = header_row.index('송장번호') + 1
    else:
        # 컬럼 추가
        waybill_idx = ws.max_column + 1
        ws.cell(1, waybill_idx, WAYBILL_COL_HEADER)

    # 매칭 진행
    used_waybills: Dict[Tuple[str, str], int] = {}  # 키별 사용 카운터
    filled = 0
    unmatched = []
    duplicates = []

    for r in range(2, ws.max_row + 1):
        name = str(ws.cell(r, name_idx).value or '').strip()
        phone = _normalize_phone(ws.cell(r, phone1_idx).value)
        if not name:
            continue
        key = (name, phone)
        candidates = waybill_map.get(key)
        if not candidates:
            unmatched.append({'행': r, '수령인명': name, '연락처': phone})
            continue
        idx = used_waybills.get(key, 0)
        if idx >= len(candidates):
            # 다원 채번에 동일 (이름,전화) 키가 부족 — 매칭 실패
            unmatched.append({'행': r, '수령인명': name, '연락처': phone,
                              '원인': '동일 키 채번 부족'})
            continue
        waybill = candidates[idx]
        used_waybills[key] = idx + 1
        ws.cell(r, waybill_idx, waybill)
        filled += 1
        if len(candidates) > 1:
            duplicates.append({
                '수령인명': name, '연락처': phone,
                '키 행': r, '사용한 채번': waybill, '전체 채번 수': len(candidates),
            })

    # 다원 채번에 있으나 메이커스에 매칭 안 된 잔여
    leftover = []
    for key, candidates in waybill_map.items():
        used = used_waybills.get(key, 0)
        if used < len(candidates):
            for wb_no in candidates[used:]:
                leftover.append({'수취인': key[0], '전화': key[1], '운송장번호': wb_no})

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), {
        'filled': filled,
        'duplicates': duplicates,
        'unmatched': unmatched,
        'leftover_waybills': leftover,
    }

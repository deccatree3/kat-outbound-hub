"""
이지어드민(EZA) 발주서 / 송장 빌더.

채널별 raw → EZA 업로드 양식. EZA 안에서 다른 채널과 통합되어 다원으로.

현재 지원:
  - 메이커스 (cachers_makers) → 메이커스 EZA 발주서 (8컬럼, .xls)
  - 국내몰 송장 — 다원 채번 → EZA 송장 업로드 양식 (.xlsx)
"""
import datetime
import io
from typing import Dict, List

import openpyxl
import xlrd
import xlwt


# ─── 메이커스 EZA 발주서 (8컬럼) ─────────────────────────────────────

MAKERS_EZA_HEADERS = [
    '주문번호', '상품명', '수량', '주문일',
    '수령인', '수령자연락처', '주소', '배송메모',
]


def _to_excel_date(value) -> tuple:
    """주문일시 → (날짜셀 값, 적용 스타일).
    datetime/문자열 → datetime.date 객체. 실패 시 (원본 문자열, None).
    """
    if isinstance(value, datetime.datetime):
        return value.date(), 'date'
    if isinstance(value, datetime.date):
        return value, 'date'
    if isinstance(value, str) and value:
        try:
            return datetime.datetime.strptime(value[:19], '%Y-%m-%d %H:%M:%S').date(), 'date'
        except ValueError:
            try:
                return datetime.datetime.strptime(value[:10], '%Y-%m-%d').date(), 'date'
            except ValueError:
                return value, None
    return ('' if value is None else str(value)), None


# ─── 국내몰 송장 양식 (이지어드민 업로드용) ──────────────────────────────

EZA_WAYBILL_DEFAULT_CARRIER = 'CJ대한통운'


def _normalize_waybill(value) -> str:
    if value is None:
        return ''
    return ''.join(c for c in str(value) if c.isdigit())


def _normalize_order_no(value) -> str:
    """주문번호 정규화. 숫자 셀은 int 변환 (12345.0→'12345'),
    텍스트 셀은 그대로 유지 (예: '130755679_1' — Python float() 가 underscore 를
    천단위 구분자로 해석하지 않도록 string 변환 시 float 시도 금지)."""
    if value in ('', None):
        return ''
    if isinstance(value, (int, float)):
        # xlrd 숫자 셀 (12345.0 → '12345')
        try:
            return str(int(value))
        except (ValueError, TypeError):
            return str(value)
    # 텍스트 셀 — 변환 시도 없이 그대로 유지
    return str(value).strip()


def parse_daone_invoice_xls(data: bytes,
                            default_carrier: str = EZA_WAYBILL_DEFAULT_CARRIER
                            ) -> tuple[List[tuple], List[Dict]]:
    """다원 채번.xls → [(carrier, waybill, order_no), ...] + skip 리스트.
    다원 양식엔 택배사 컬럼이 없어 default_carrier 적용.
    """
    wb_in = xlrd.open_workbook(file_contents=data)
    ws_in = wb_in.sheet_by_index(0)
    if ws_in.nrows < 2:
        return [], [{'원인': '데이터 없음'}]
    headers = [str(ws_in.cell_value(0, c)).strip() for c in range(ws_in.ncols)]

    def find_idx(*names):
        for n in names:
            if n in headers:
                return headers.index(n)
        return None

    # '주문번호' (일반판매 다원 채번) / '출하의뢰항번' (로켓그로스 택배 = '{itr_id}_{box_no}')
    order_i = find_idx('주문번호', '출하의뢰항번')
    waybill_i = find_idx('운송장번호', '송장번호')
    if order_i is None or waybill_i is None:
        raise RuntimeError(
            f"다원 채번 파일에서 '주문번호'/'출하의뢰항번' 또는 '운송장번호' 컬럼을 찾지 못함. "
            f"실제 헤더: {headers}"
        )

    triples = []
    skipped = []
    for r in range(1, ws_in.nrows):
        order_no = _normalize_order_no(ws_in.cell_value(r, order_i))
        waybill = _normalize_waybill(ws_in.cell_value(r, waybill_i))
        if not order_no or not waybill:
            skipped.append({'source': '다원', '행': r + 1,
                            '주문번호': order_no, '송장번호': waybill,
                            '원인': '필수값 없음'})
            continue
        triples.append((default_carrier, waybill, order_no))
    return triples, skipped


def parse_3pl_invoice_xlsx(data: bytes,
                           default_carrier: str = EZA_WAYBILL_DEFAULT_CARRIER
                           ) -> tuple[List[tuple], List[Dict]]:
    """3PL 출고요청서.xlsx (우리 25컬럼 양식) — 송장번호 채워진 상태.
    택배사는 무조건 default_carrier (CJ대한통운) — 양식 내 '택배사' 컬럼은 무시.
    """
    wb_in = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws_in = wb_in.active
    if ws_in.max_row < 2:
        return [], [{'원인': '데이터 없음'}]
    headers = [str(ws_in.cell(1, c).value or '').strip()
               for c in range(1, ws_in.max_column + 1)]

    def find_idx(*names):
        for n in names:
            if n in headers:
                return headers.index(n) + 1   # openpyxl 1-indexed
        return None

    order_i = find_idx('주문번호', '고객주문번호')
    waybill_i = find_idx('송장번호', '운송장번호')
    if order_i is None or waybill_i is None:
        raise RuntimeError(
            f"3PL 파일에서 '주문번호' 또는 '송장번호' 컬럼을 찾지 못함. "
            f"실제 헤더: {headers}"
        )

    triples = []
    skipped = []
    for r in range(2, ws_in.max_row + 1):
        order_no = _normalize_order_no(ws_in.cell(r, order_i).value)
        waybill = _normalize_waybill(ws_in.cell(r, waybill_i).value)
        if not order_no or not waybill:
            skipped.append({'source': '3PL', '행': r,
                            '주문번호': order_no, '송장번호': waybill,
                            '원인': '필수값 없음'})
            continue
        triples.append((default_carrier, waybill, order_no))
    return triples, skipped


def build_eza_waybill_from_triples(triples: List[tuple]) -> bytes:
    """[(carrier, waybill, order_no), ...] → 이지어드민 송장 양식.xlsx.

    출력 컬럼 위치:
      A 택배사 / D 송장번호 / E 관리번호
    """
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = 'Sheet1'
    ws_out.cell(1, 1, '택배사')
    ws_out.cell(1, 4, '송장번호')
    ws_out.cell(1, 5, '주문번호')
    for letter, w in [('A', 13), ('B', 13), ('C', 13), ('D', 15), ('E', 13)]:
        ws_out.column_dimensions[letter].width = w

    for i, (carrier, waybill, order_no) in enumerate(triples, 2):
        ws_out.cell(i, 1, carrier)
        ws_out.cell(i, 4, waybill)
        ws_out.cell(i, 5, order_no)

    buf = io.BytesIO()
    wb_out.save(buf)
    return buf.getvalue()


def build_eza_waybill_xlsx(daone_invoice_xls_bytes: bytes,
                           carrier: str = EZA_WAYBILL_DEFAULT_CARRIER) -> tuple[bytes, Dict]:
    """다원 채번.xls 단일 파일 → 이지어드민 송장 양식.xlsx (하위 호환)."""
    triples, skipped = parse_daone_invoice_xls(daone_invoice_xls_bytes, carrier)
    return build_eza_waybill_from_triples(triples), {
        'filled': len(triples), 'skipped': skipped,
    }


def build_makers_eza_xls(makers_rows: List[Dict]) -> bytes:
    """메이커스 주문내역 dict → EZA 발주서 .xls bytes.

    매핑 미사용 — 상품명 + 옵션 그대로 한 컬럼으로 결합. EZA가 자체 매핑 보유.
    """
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Sheet1')

    date_style = xlwt.easyxf(num_format_str='YYYY-MM-DD')

    for ci, h in enumerate(MAKERS_EZA_HEADERS):
        ws.write(0, ci, h)

    for ri, r in enumerate(makers_rows, 1):
        order_no = r.get('주문번호', '')
        # 주문번호: 숫자형이면 정수로
        try:
            order_no_v = int(order_no) if str(order_no).strip() else ''
        except (ValueError, TypeError):
            order_no_v = str(order_no)

        product = (r.get('상품') or '').strip()
        option = (r.get('옵션') or '').strip()
        product_full = f"{product}_{option}" if option else product

        try:
            qty = int(float(r.get('수량', 1) or 1))
        except (ValueError, TypeError):
            qty = 1

        date_val, date_kind = _to_excel_date(r.get('주문일시'))

        recipient = (r.get('수령인명') or '').strip()
        phone = str(r.get('수령인 연락처1', '')).strip()
        address = (r.get('배송주소') or '').strip()
        memo = (r.get('배송메시지') or '').strip()

        ws.write(ri, 0, order_no_v)
        ws.write(ri, 1, product_full)
        ws.write(ri, 2, qty)
        if date_kind == 'date':
            ws.write(ri, 3, date_val, date_style)
        else:
            ws.write(ri, 3, date_val)
        ws.write(ri, 4, recipient)
        ws.write(ri, 5, phone)
        ws.write(ri, 6, address)
        ws.write(ri, 7, memo)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

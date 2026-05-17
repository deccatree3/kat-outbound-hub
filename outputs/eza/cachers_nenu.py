"""캐처스가 판매하는 네뉴 상품(매입리스트) 품절 홀딩 로직.

판매처는 캐처스지만 실제 화주가 네뉴인 품목군이 있다. 이 상품이 캐처스 재고
기준 품절/부족이면, 해당 상품이 든 **합포장(같은 수취인) 주문 전체**를 이번 차수
다원 발주서에서 제외하고, 네뉴→캐처스 재고이동용 이지어드민 발주서를 만든다.
(상품만 빼고 출고 후 따로 보내면 배송비가 이중 발생하므로 묶음 전체 홀딩.)

이 모듈은 순수 로직 + 매입리스트 로더만 담당한다. DB(box입수) 조회와 UI 는
channels/domestic/page.py 가 수행한다.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

# 레포 번들 경로 (배포 환경에서도 읽히도록 코드와 함께 커밋됨).
# 매입리스트 갱신 시 이 파일만 교체 후 커밋.
PURCHASE_LIST_PATH = Path(__file__).with_name("cachers_nenu_purchase_list.xlsx")
PURCHASE_LIST_SHEET = "캐처스_WMS"
PURCHASE_LIST_HEADER_ROW = 7  # 1-indexed: R7=헤더, R8~ 데이터

# 합포장(묶음배송) 그룹 키 — 다원/출고요청서 컬럼 기준 (사용자 확정)
GROUP_KEY_COLS = (
    "수취인명", "수취인연락처1", "수취인연락처2", "수취인우편번호", "수취인주소1",
)

STATUS_MOVE = "이동필요"   # 완전품절 또는 주문량 부족
STATUS_WATCH = "관찰"      # 주문 차감 후 잔여 < box입수 50%


@dataclass(frozen=True)
class PurchaseItem:
    code: str          # 캐처스 품목코드 (EZA 상품메모/제품코드 · 재고파일 품목코드와 동일 키)
    name: str          # 품목명
    barcode: str        # 다원 품목코드(WMS 바코드) — box입수 조인 키


@dataclass
class AffectedProduct:
    code: str
    name: str
    barcode: str
    ordered: int           # 이번 차수 해당상품 주문수량 합
    available: int         # 캐처스 가용재고 (가능수량 합, RELEASEAREA 제외, 미등장=0)
    box_qty: int | None    # box입수 (WmsProduct.box_qty), 없으면 None
    status: str            # STATUS_MOVE | STATUS_WATCH


def load_purchase_list(path: str | Path = PURCHASE_LIST_PATH) -> list[PurchaseItem]:
    """매입리스트(캐처스_WMS 시트) → PurchaseItem 리스트.

    R7 헤더(캐처스 품목코드 / 품목명 / 다원 품목코드 / 비고), R8~ 데이터.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[PURCHASE_LIST_SHEET]
    items: list[PurchaseItem] = []
    for r in range(PURCHASE_LIST_HEADER_ROW + 1, ws.max_row + 1):
        code = ws.cell(r, 1).value
        if code is None or str(code).strip() == "":
            continue
        items.append(PurchaseItem(
            code=str(code).strip(),
            name=str(ws.cell(r, 2).value or "").strip(),
            barcode=str(ws.cell(r, 3).value or "").strip(),
        ))
    return items


def group_key(row: dict) -> tuple:
    """합포장 그룹 키 (수취인명+연락처1+연락처2+우편번호+주소1)."""
    return tuple(str(row.get(c, "") or "").strip() for c in GROUP_KEY_COLS)


def _row_code(row: dict) -> str:
    """다원 행의 캐처스 품목코드 = 제품코드 (EZA 상품메모 폴백이 이미 반영됨)."""
    return str(row.get("제품코드", "") or "").strip()


def compute_affected_products(
    daone_rows: list[dict],
    stock_agg: dict[str, dict[str, Any]],
    purchase_list: list[PurchaseItem],
    box_qty_by_code: dict[str, int | None],
) -> list[AffectedProduct]:
    """이번 차수 캐처스 주문 중 매입리스트 상품의 품절/부족/관찰 판정.

    - ordered  = 해당 제품코드 daone 행들의 주문수량 합
    - available = stock_agg[code].available_qty (미등장 → 0)
    - 이동필요 = available <= 0  OR  available < ordered
    - 관찰    = (이동필요 아님) AND box_qty 존재 AND (available-ordered) < box_qty*0.5
    - 그 외(정상)는 결과에서 제외
    반환은 status 우선(이동필요 먼저) → 품목명 순.
    """
    by_code = {p.code: p for p in purchase_list}

    ordered: dict[str, int] = {}
    for row in daone_rows:
        code = _row_code(row)
        if code not in by_code:
            continue
        try:
            q = int(float(row.get("주문수량", 0) or 0))
        except (ValueError, TypeError):
            q = 0
        ordered[code] = ordered.get(code, 0) + q

    out: list[AffectedProduct] = []
    for code, qty in ordered.items():
        item = by_code[code]
        avail = int(stock_agg.get(code, {}).get("available_qty", 0) or 0)
        bq = box_qty_by_code.get(code)
        if avail <= 0 or avail < qty:
            status = STATUS_MOVE
        elif bq and (avail - qty) < bq * 0.5:
            status = STATUS_WATCH
        else:
            continue
        out.append(AffectedProduct(
            code=code, name=item.name, barcode=item.barcode,
            ordered=qty, available=avail, box_qty=bq, status=status,
        ))

    out.sort(key=lambda a: (a.status != STATUS_MOVE, a.name))
    return out


def split_held_orders(
    daone_rows: list[dict],
    held_codes: set[str],
) -> tuple[list[dict], list[dict], int]:
    """선택된 품절코드가 든 합포장(수취인) 그룹 전체를 분리.

    반환: (출고분 daone_rows, 홀딩분 daone_rows, 홀딩 그룹 수).
    홀딩 그룹 = 그룹 내 행 중 하나라도 제품코드 ∈ held_codes 인 그룹.
    """
    if not held_codes:
        return list(daone_rows), [], 0

    held_groups: set[tuple] = set()
    for row in daone_rows:
        if _row_code(row) in held_codes:
            held_groups.add(group_key(row))

    shipped, held = [], []
    for row in daone_rows:
        (held if group_key(row) in held_groups else shipped).append(row)
    return shipped, held, len(held_groups)

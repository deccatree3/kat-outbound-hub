"""업로드 파일 자동 분류: 파일 타입 + 업체 식별.

파일 타입 판별 (파일명 기반):
  - inventory_health_sku_info_* → 'coupang_inventory'
  - Document_* → 'wms_inventory'
  - generated_excel* → 'coupang_template'
  - 쿠팡 재고이동건_* / *재고이동* → 'movement'

업체 식별 (내용 기반):
  - 옵션ID/바코드를 추출 → DB의 coupang_product/wms_product 에서 company_name 조회
"""
from __future__ import annotations

from collections import Counter
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy import select

from rocketgrowth.db import get_session
from rocketgrowth.models import CoupangProduct, WmsProduct


FILE_TYPE_COUPANG = "coupang_inventory"
FILE_TYPE_WMS = "wms_inventory"
FILE_TYPE_TEMPLATE = "coupang_template"
FILE_TYPE_MOVEMENT = "movement"
FILE_TYPE_UNKNOWN = "unknown"

FILE_TYPE_LABELS = {
    FILE_TYPE_COUPANG: "쿠팡 재고",
    FILE_TYPE_WMS: "WMS 재고",
    FILE_TYPE_TEMPLATE: "쿠팡 입고생성",
    FILE_TYPE_MOVEMENT: "재고이동",
    FILE_TYPE_UNKNOWN: "미분류",
}


@dataclass
class ClassifiedFile:
    file: Any               # Streamlit UploadedFile
    file_type: str          # FILE_TYPE_*
    company: str | None     # 식별된 업체명 (None=미식별)
    confidence: float = 0.0 # 0~1


@dataclass
class CompanyFileGroup:
    company: str
    files: dict[str, Any] = field(default_factory=dict)  # {file_type: UploadedFile}

    @property
    def is_complete(self) -> bool:
        return all(
            ft in self.files
            for ft in [FILE_TYPE_COUPANG, FILE_TYPE_WMS, FILE_TYPE_TEMPLATE, FILE_TYPE_MOVEMENT]
        )

    @property
    def missing_types(self) -> list[str]:
        return [
            ft for ft in [FILE_TYPE_COUPANG, FILE_TYPE_WMS, FILE_TYPE_TEMPLATE, FILE_TYPE_MOVEMENT]
            if ft not in self.files
        ]


def classify_file_type(filename: str) -> str:
    """파일명으로 타입 판별."""
    name = filename.lower()
    if "inventory_health" in name or "sku_info" in name:
        return FILE_TYPE_COUPANG
    if name.startswith("document") or "document_" in name:
        return FILE_TYPE_WMS
    if "generated_excel" in name:
        return FILE_TYPE_TEMPLATE
    if "재고이동" in filename:
        return FILE_TYPE_MOVEMENT
    return FILE_TYPE_UNKNOWN


def identify_company_from_coupang_file(file_bytes: bytes) -> str | None:
    """쿠팡 재고현황 파일에서 옵션ID/등록상품ID/SKU ID 추출 → DB에서 업체명 조회.

    샘플 50행 + 3개 키 (옵션ID/등록상품ID/SKU ID) 모두 사용해서 매칭 다양화.
    소량만 DB 에 등록되었거나 일부 SKU 가 default '서현' 으로 잘못 저장된 경우에도
    다수결로 정확한 업체를 식별.
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        option_ids: list[int] = []
        product_ids: list[int] = []
        sku_ids: list[int] = []
        for r in range(3, min(53, ws.max_row + 1)):  # 3~52 행 = 50개 샘플
            cells = next(ws.iter_rows(min_row=r, max_row=r, values_only=True), None)
            if not cells:
                continue
            # col 2 (idx 1) = 등록상품 ID, col 3 (idx 2) = 옵션 ID, col 4 (idx 3) = SKU ID
            for ids_list, idx in ((product_ids, 1), (option_ids, 2), (sku_ids, 3)):
                v = cells[idx] if len(cells) > idx else None
                if v not in (None, "", "-"):
                    try:
                        ids_list.append(int(v))
                    except (ValueError, TypeError):
                        pass
        wb.close()
        if not (option_ids or product_ids or sku_ids):
            return None
        return _lookup_company_by_coupang_ids(
            option_ids=option_ids, product_ids=product_ids, sku_ids=sku_ids,
        )
    except Exception:
        return None


def identify_company_from_wms_file(file_bytes: bytes) -> str | None:
    """WMS Document 파일에서 바코드 추출 → DB에서 업체명 조회. 샘플 50행."""
    try:
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_bytes)
        ws = wb.sheet_by_index(0)
        barcodes = []
        for r in range(1, min(51, ws.nrows)):
            val = ws.row_values(r)[0] if ws.ncols > 0 else None
            if val:
                barcodes.append(str(val).strip())
        if not barcodes:
            return None
        return _lookup_company_by_barcodes(barcodes)
    except Exception:
        # xlsx 형식일 수도
        try:
            wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
            ws = wb.active
            barcodes = []
            for r in range(2, min(52, ws.max_row + 1)):
                val = ws.cell(row=r, column=1).value
                if val:
                    barcodes.append(str(val).strip())
            wb.close()
            if not barcodes:
                return None
            return _lookup_company_by_barcodes(barcodes)
        except Exception:
            return None


def identify_company_from_template(file_bytes: bytes) -> str | None:
    """쿠팡 업로드 양식에서 옵션ID 추출 → 업체 식별.

    시트는 '로켓그로스 입고' 우선, 없으면 active. 헤더 1~4행, 데이터 5행부터.
    옵션 ID 열은 G열(7) 기본이며, 헤더에 '옵션 ID' 가 있으면 그 위치로 보정.
    데이터 5~50 행 스캔 (샘플 50개).
    """
    # read_only=True 가 일부 generated_excel 에서 dimensions 를 1x1 로 잘못 보고하는 버그가 있어
    # full mode 로 읽는다 (파일 크기 작아 비용 무시할 수준).
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        ws = wb["로켓그로스 입고"] if "로켓그로스 입고" in wb.sheetnames else wb.active

        # 헤더에서 '옵션 ID' 컬럼 위치 탐색 (못 찾으면 7=G 기본값)
        opt_col = 7
        for r in range(1, min(5, ws.max_row + 1)):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v and "옵션 ID" in str(v):
                    opt_col = c
                    break
            else:
                continue
            break

        option_ids = []
        for r in range(5, min(55, ws.max_row + 1)):  # 5~54 = 50개 샘플
            val = ws.cell(row=r, column=opt_col).value
            if val in (None, ""):
                continue
            try:
                option_ids.append(int(str(val).strip()))
            except (ValueError, TypeError):
                pass
        wb.close()
        if not option_ids:
            return None
        return _lookup_company_by_coupang_ids(option_ids=option_ids)
    except Exception:
        return None


def identify_company_from_movement(file_bytes: bytes) -> str | None:
    """재고이동건 파일 시트명에서 업체명 추출, 또는 바코드로 식별."""
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True)
        for sname in wb.sheetnames:
            if sname == "form":
                continue
            # 시트명 형식: "MMDD(서현, 밀크런, FC)"
            if "(" in sname and "," in sname:
                parts = sname.split("(")[1].split(",")
                company = parts[0].strip()
                wb.close()
                return company
        wb.close()
        return None
    except Exception:
        return None


def _lookup_company_by_option_ids(option_ids: list[int]) -> str | None:
    """DB에서 옵션ID로 업체명 조회. 다수결. (구 호환용 — 신규는 _lookup_company_by_coupang_ids)."""
    return _lookup_company_by_coupang_ids(option_ids=option_ids)


def _lookup_company_by_coupang_ids(
    option_ids: list[int] | None = None,
    product_ids: list[int] | None = None,
    sku_ids: list[int] | None = None,
) -> str | None:
    """3개 키 (옵션ID / 등록상품ID / SKU ID) 합쳐서 DB에서 업체명 조회. 다수결.

    한 키만으로 매칭 실패해도 다른 키로 보완. 더 많은 행이 매칭되면 다수결 정확도 ↑.
    """
    option_ids = option_ids or []
    product_ids = product_ids or []
    sku_ids = sku_ids or []
    if not (option_ids or product_ids or sku_ids):
        return None
    rows: list[str] = []
    with get_session() as session:
        if option_ids:
            rows.extend(session.execute(
                select(CoupangProduct.company_name).where(
                    CoupangProduct.coupang_option_id.in_(option_ids)
                )
            ).scalars().all())
        if product_ids:
            rows.extend(session.execute(
                select(CoupangProduct.company_name).where(
                    CoupangProduct.coupang_product_id.in_(product_ids)
                )
            ).scalars().all())
        if sku_ids:
            rows.extend(session.execute(
                select(CoupangProduct.company_name).where(
                    CoupangProduct.sku_id.in_(sku_ids)
                )
            ).scalars().all())
    if not rows:
        return None
    counter = Counter(rows)
    return counter.most_common(1)[0][0]


def _lookup_company_by_barcodes(barcodes: list[str]) -> str | None:
    """DB에서 WMS 바코드로 업체명 조회. 다수결."""
    with get_session() as session:
        rows = session.execute(
            select(WmsProduct.company_name).where(
                WmsProduct.wms_barcode.in_(barcodes)
            )
        ).scalars().all()
    if not rows:
        return None
    counter = Counter(rows)
    return counter.most_common(1)[0][0]


def _get_known_companies() -> set[str]:
    """DB에 등록된 업체명 목록."""
    try:
        with get_session() as session:
            cp = set(session.execute(select(CoupangProduct.company_name).distinct()).scalars().all())
            wms = set(session.execute(select(WmsProduct.company_name).distinct()).scalars().all())
            return cp | wms
    except Exception:
        return set()


def identify_company_from_filename(filename: str, known_companies: set[str]) -> str | None:
    """파일명에 업체명이 포함되어 있으면 반환. 가장 먼저 시도."""
    for company in known_companies:
        if company in filename:
            return company
    return None


def classify_uploaded_files(uploaded_files: list) -> tuple[list[ClassifiedFile], dict[str, CompanyFileGroup]]:
    """업로드된 파일들을 분류하고 업체별로 그룹핑.

    식별 우선순위:
      1. 파일명에 업체명 포함 (가장 빠르고 확실)
      2. 파일 내용에서 옵션ID/바코드 → DB 조회

    Returns:
        (classified_list, company_groups)
    """
    known = _get_known_companies()
    classified: list[ClassifiedFile] = []

    for f in uploaded_files:
        ftype = classify_file_type(f.name)
        company = None

        # 1순위: 파일명에서 업체명 식별
        company = identify_company_from_filename(f.name, known)

        # 2순위: 파일 내용에서 식별
        if not company:
            file_bytes = f.getvalue()
            if ftype == FILE_TYPE_COUPANG:
                company = identify_company_from_coupang_file(file_bytes)
            elif ftype == FILE_TYPE_WMS:
                company = identify_company_from_wms_file(file_bytes)
            elif ftype == FILE_TYPE_TEMPLATE:
                company = identify_company_from_template(file_bytes)
            elif ftype == FILE_TYPE_MOVEMENT:
                company = identify_company_from_movement(file_bytes)

        classified.append(ClassifiedFile(
            file=f,
            file_type=ftype,
            company=company,
            confidence=1.0 if company else 0.0,
        ))

    # 업체별 그룹핑
    groups: dict[str, CompanyFileGroup] = {}
    for cf in classified:
        if not cf.company:
            continue
        if cf.company not in groups:
            groups[cf.company] = CompanyFileGroup(company=cf.company)
        groups[cf.company].files[cf.file_type] = cf.file

    return classified, groups

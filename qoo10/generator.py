"""
QSM detail.csv → KSE OMS Outbound 양식 변환.
QSM brief.csv + 송장번호 매핑 → QSM 업로드용 CSV 생성.
"""
import csv
import copy
import io
import os
import sys
import datetime
from typing import List, Dict, Tuple

import openpyxl
from openpyxl.utils import get_column_letter

_THIS = os.path.dirname(os.path.abspath(__file__))
OUTBOUND_TEMPLATE = os.path.join(_THIS, "templates", "outbound_template.xlsx")

# db/pg.py import
_THIS = os.path.dirname(os.path.abspath(__file__))
_BASE = os.path.dirname(_THIS)
if os.path.join(_BASE, "db") not in sys.path:
    sys.path.insert(0, os.path.join(_BASE, "db"))
import pg

# Outbound 51컬럼 순서 (헤더는 2줄 합침: 日本語\n영문코드)
OUTBOUND_HEADERS = [
    ("倉庫コード", "CTKEY"),
    ("荷主コード", "OWKEY"),
    ("出庫予定日", "OR_HDDATE"),
    ("注文日", "OR_DATE"),
    ("注文タイプ", "ORHDTYPE"),
    ("商品コード", "ICMPKEY"),
    ("商品オプション名称", "IC_OPTION"),
    ("商品単位コード", "ICUTKEY"),
    ("代替コード", "SBKEY"),
    ("物流グループコード", "LOGGRPCD"),
    ("販売先コード", "STORE_KEY"),
    ("単位", "UOM"),
    ("予定数量", "EXQTY"),
    ("生産日", "PRODUCTDATE"),
    ("有効日", "EXPIREDATE"),
    ("注文番号", "EXTERNORDERKEY"),
    ("仕入先コード", "ACKEY"),
    ("仕入先名/受取人名", "ACNAME"),
    ("電話番号", "TEL"),
    ("携帯電話番号", "CP"),
    ("FAX番号", "FAX"),
    ("担当者", "CONTACT"),
    ("国コード", "COUNTRYCODE"),
    ("郵便番号コード", "ZCKEY"),
    ("基本住所", "ADDRESS1"),
    ("詳細住所", "ADDRESS2"),
    ("都市", "CITY"),
    ("都道府県(州)", "STATE"),
    ("配送会社", "DLCOMPANY"),
    ("注文配送運賃タイプ", "ODPAYTYPE"),
    ("配達指定日", "DLDATE"),
    ("配達時間帯", "DLTIME"),
    ("注文担当者", "OR_USER_ID"),
    ("注文先名", "ORNAME"),
    ("注文先電話番号", "ORTEL"),
    ("注文先FAX番号", "ORFAX"),
    ("注文先担当者", "ORCONTACT"),
    ("注文先国コード", "ORCOUNTRYCODE"),
    ("注文先郵便番号", "ORZCKEY"),
    ("注文先基本住所", "ORADDRESS1"),
    ("注文先詳細住所", "ORADDRESS2"),
    ("注文先都市", "ORCITY"),
    ("注文先都道府県(州)", "ORSTATE"),
    ("単位原価", "COSTPRICE"),
    ("販売価格", "SALEPRICE"),
    ("ベンダーコード", "VDKEY"),
    ("集合梱包情報", "PACKAGESOURCE"),
    ("コメント1", "COMMENTS1"),
    ("コメント2", "COMMENTS2"),
    ("TC/DC", "ATTRIBUTE1"),
    ("一般/保税", "ATTRIBUTE2"),
]


def load_kse_sku_catalog(location: str = 'JP') -> List[Dict]:
    """SKU 카탈로그 (전체). 매핑 테이블에서 distinct 추출.

    location 인자는 하위 호환용 (현재 사용 안 함).
    창고 구분이 필요하면 channel_product_mapping(channel) 으로 도출.

    우선순위:
      1. channel_product_mapping (모든 매핑의 sku_codes/item_codes)
      2. stock_snapshots + shipments UNION (자매 프로젝트 fallback, 비어있을 때만)
      3. qoo10_outbound 이력 fallback
    """
    try:
        from db import mapping as _m
        rows = _m.list_known_skus()
        if rows:
            return rows
    except Exception:
        pass

    try:
        conn = pg.connect(autocommit=True)
        with conn.cursor() as cur:
            cur.execute("""
                SELECT DISTINCT sku_code, sku_name FROM (
                    SELECT sku_code, sku_name FROM stock_snapshots
                    UNION
                    SELECT sku_code, sku_name FROM shipments
                ) t
                WHERE sku_code IS NOT NULL AND sku_code != ''
                  AND sku_name IS NOT NULL
                ORDER BY sku_name
            """)
            rows = cur.fetchall()
        conn.close()
        return [{'sku_code': r[0], 'sku_name': r[1]} for r in rows]
    except Exception:
        pass
    try:
        conn = pg.connect(autocommit=True)
        with conn.cursor() as cur:
            cur.execute("""
                SELECT DISTINCT sku_code, sku_name FROM qoo10_outbound
                WHERE sku_code IS NOT NULL AND sku_code != ''
                  AND sku_name IS NOT NULL AND sku_name != ''
                ORDER BY sku_name
            """)
            rows = cur.fetchall()
        conn.close()
        return [{'sku_code': r[0], 'sku_name': r[1]} for r in rows]
    except Exception:
        return []


CHANNEL_QOO10_JAPAN = 'qoo10_japan'


def add_mapping(qoo10_name: str, qoo10_option: str,
                skus: List[Tuple[str, str, int]], enabled: bool = True):
    """Qoo10 일본 매핑 upsert. skus = [(sku_code, sku_name, qty), ...]
    enabled 인자는 하위 호환용 (채널 분리로 의미 잃음).
    """
    from db import mapping as _m
    _m.upsert(CHANNEL_QOO10_JAPAN, qoo10_name, qoo10_option, skus)


def delete_mapping(qoo10_name: str, qoo10_option: str):
    """Qoo10 일본 매핑 삭제"""
    from db import mapping as _m
    _m.delete(CHANNEL_QOO10_JAPAN, qoo10_name, qoo10_option)


def load_mappings() -> Dict[Tuple[str, str], Dict]:
    """Qoo10 일본 매핑 로드. key=(상품명, 옵션)"""
    from db import mapping as _m
    return _m.load_for_channel(CHANNEL_QOO10_JAPAN)


def parse_qsm_csv(content: bytes) -> List[Dict]:
    """QSM detail.csv bytes → list of dict"""
    text = content.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    return list(reader)


def normalize_postal(code: str) -> str:
    """QSM은 '289-1733 형식 → 289-1733"""
    if not code:
        return ''
    return code.lstrip("'").strip()


HYPHEN_LIKE = {
    0x2010: 'HYPHEN',
    0x2011: 'NON-BREAKING HYPHEN',
    0x2012: 'FIGURE DASH',
    0x2013: 'EN DASH',
    0x2014: 'EM DASH',
    0x2015: 'HORIZONTAL BAR',
}


def _char_reason(ch: str) -> str:
    """제거/치환된 문자의 의미 반환 (사람이 읽을 수 있는 설명)."""
    cp = ord(ch)
    if cp in HYPHEN_LIKE:
        return f"'{ch}'({HYPHEN_LIKE[cp]}) → '-' 치환"
    if 8192 <= cp <= 8303:
        # 자주 나오는 것들 친숙한 이름 부여
        names = {
            0x2003: 'EM SPACE',
            0x2009: 'THIN SPACE',
            0x200B: 'ZERO-WIDTH SPACE',
            0x2018: 'LEFT SINGLE QUOTE',
            0x2019: 'RIGHT SINGLE QUOTE',
            0x201C: 'LEFT DOUBLE QUOTE',
            0x201D: 'RIGHT DOUBLE QUOTE',
            0x2026: 'ELLIPSIS (…)',
        }
        name = names.get(cp, f'특수 문장부호 U+{cp:04X}')
        return f"'{ch}'({name}) 제거"
    if 9728 <= cp <= 9983:
        return f"'{ch}'(기호/도형 U+{cp:04X}) 제거"
    return f"'{ch}' 변경"


def clean_special_chars(text: str) -> Tuple[str, List[str]]:
    """
    템플릿 VBA Module1.CleanSpecialChars 포팅 + 주소 보호 개선.

    삭제 범위:
      - U+2000~U+206F: 일반 문장 부호
      - U+2600~U+26FF: 기호 및 도형

    예외 (주소 왜곡 방지):
      - U+2010~U+2015 하이픈/대시 계열은 일반 '-'로 치환

    반환: (정제된 텍스트, 사유 리스트)
    """
    if not text:
        return '', []
    reasons = []
    out_chars = []
    for ch in text:
        cp = ord(ch)
        if cp in HYPHEN_LIKE:
            out_chars.append('-')
            reasons.append(_char_reason(ch))
        elif 8192 <= cp <= 8303 or 9728 <= cp <= 9983:
            reasons.append(_char_reason(ch))
            # 삭제 (문자 추가 안 함)
        else:
            out_chars.append(ch)
    return ''.join(out_chars), reasons


def normalize_order_date(qsm_date: str) -> str:
    """2026/04/15 19:12:16 → 20260415"""
    if not qsm_date:
        return ''
    try:
        dt = datetime.datetime.strptime(qsm_date.strip(), '%Y/%m/%d %H:%M:%S')
        return dt.strftime('%Y%m%d')
    except ValueError:
        # 이미 YYYYMMDD일 수 있음
        digits = ''.join(c for c in qsm_date if c.isdigit())
        return digits[:8] if len(digits) >= 8 else qsm_date


def generate_outbound_rows(qsm_rows: List[Dict], mappings: Dict) -> Tuple[List[Dict], List[Dict], List[Dict]]:
    """
    QSM detail 행들 → Outbound 행들 변환.
    Power Query 로직 준수:
      1. 취합대상(enabled)=y만 필터
      2. 품목코드("SKU1,qty1,SKU2,qty2") split & expand → N SKU 행
      3. 予定数量 = QSM수량 × 매핑 SKU당수량
      4. 정렬: 장바구니번호 ASC, 주문번호 ASC, 품목(SKU) ASC
      5. 注文番号는 장바구니번호 사용 (같은 장바구니 = 합포장)
      6. 주소는 VBA CleanSpecialChars 적용 (★ ◆ 등 제거)
    반환: (출고 행들, 미매핑/에러 행들, 주소 정제 변경 이력)
    """
    today = datetime.date.today().strftime('%Y%m%d')
    outbound_rows = []
    errors = []
    addr_changes = []  # 정제로 변경된 주소 목록 (사용자 확인용)

    for q in qsm_rows:
        name = (q.get('상품명') or '').strip()
        option = (q.get('옵션정보') or '').strip()
        qsm_qty = int(q.get('수량', '1') or 1)
        cart_no = (q.get('장바구니번호') or '').strip()
        order_no = (q.get('주문번호') or '').strip()

        m = mappings.get((name, option))
        if m is None:
            errors.append({
                '장바구니번호': cart_no, '주문번호': order_no,
                '상품명': name, '옵션정보': option,
                '원인': '상품 매핑 없음',
            })
            continue
        if not m['enabled']:
            errors.append({
                '장바구니번호': cart_no, '주문번호': order_no,
                '상품명': name, '옵션정보': option,
                '원인': '매핑 비활성(취급 안함)',
            })
            continue

        # 주소 특수문자 정제 (VBA CleanSpecialChars)
        orig_addr = (q.get('주소') or '').strip()
        clean_addr, clean_reasons = clean_special_chars(orig_addr)
        if orig_addr != clean_addr:
            # 중복 사유 제거 (같은 문자 여러 번이면 한 번만)
            unique_reasons = list(dict.fromkeys(clean_reasons))
            addr_changes.append({
                '장바구니번호': cart_no, '주문번호': order_no,
                '원본주소': orig_addr, '정제주소': clean_addr,
                '사유': ' / '.join(unique_reasons),
            })

        # 전화번호: 수취인핸드폰 > 수취인전화 (값이 "-"면 skip)
        tel_cands = [
            q.get('수취인핸드폰번호', '').strip(),
            q.get('수취인전화번호', '').strip(),
        ]
        tel = next((t for t in tel_cands if t and t != '-'), '')

        # SKU별 1행 생성 (세트 상품은 N행으로 분할)
        for sku_code, sku_unit_qty in zip(m['sku_codes'], m['quantities']):
            if not sku_code or sku_code == '-':
                continue
            row = {h[0]: '' for h in OUTBOUND_HEADERS}
            row['倉庫コード'] = 'KE00003'
            row['荷主コード'] = 'katchers'
            row['出庫予定日'] = today
            row['注文日'] = normalize_order_date(q.get('주문일', ''))
            row['商品コード'] = sku_code
            row['予定数量'] = sku_unit_qty * qsm_qty  # 핵심: 매핑수량 × QSM수량
            row['注文番号'] = cart_no  # 장바구니번호 사용 (합포장)
            row['仕入先名/受取人名'] = q.get('수취인명', '')
            row['電話番号'] = tel
            row['国コード'] = 'JPN'
            row['郵便番号コード'] = normalize_postal(q.get('우편번호', ''))
            row['基本住所'] = clean_addr
            row['配送会社'] = '320'  # 사가와
            row['注文配送運賃タイプ'] = '10'  # 선불
            row['注文先名'] = q.get('수취인명', '')
            row['注文先電話番号'] = tel
            row['注文先国コード'] = 'JPN'
            row['注文先郵便番号'] = normalize_postal(q.get('우편번호', ''))
            row['注文先基本住所'] = clean_addr
            # 정렬용 내부 키 (마지막에 제거됨)
            row['_sort_cart'] = cart_no
            row['_sort_order'] = order_no
            row['_sort_sku'] = sku_code
            outbound_rows.append(row)

    # Power Query와 동일한 정렬: 장바구니 ASC → 주문 ASC → SKU ASC
    outbound_rows.sort(key=lambda r: (r['_sort_cart'], r['_sort_order'], r['_sort_sku']))
    for r in outbound_rows:
        r.pop('_sort_cart', None)
        r.pop('_sort_order', None)
        r.pop('_sort_sku', None)

    return outbound_rows, errors, addr_changes


def compute_audit(qsm_rows: List[Dict], outbound_rows: List[Dict],
                  mappings: Dict) -> Dict:
    """
    표1 시트 검수 지표 (row 2~5) 계산.
    OMS 업로드 결과와 수치 비교하는 용도.
    """
    # enabled QSM 행만
    enabled_qsm = []
    for q in qsm_rows:
        name = (q.get('상품명') or '').strip()
        option = (q.get('옵션정보') or '').strip()
        m = mappings.get((name, option))
        if m and m['enabled']:
            enabled_qsm.append((q, m))

    # 1. 총 상품 수량 = SUM of mapping 수량 합 (template D열 = 수량 컬럼)
    total_item_qty = 0
    # 2. 주문 업로드 개수 = SUM of SKU 개수 (output row 수)
    upload_row_count = 0
    for _, m in enabled_qsm:
        for sku_code, sku_qty in zip(m['sku_codes'], m['quantities']):
            if sku_code and sku_code != '-':
                total_item_qty += sku_qty
                upload_row_count += 1

    # 3. 송장번호 개수 = unique 장바구니번호 (enabled QSM)
    unique_carts = len({(q.get('장바구니번호') or '').strip() for q, _ in enabled_qsm
                        if (q.get('장바구니번호') or '').strip()})
    # 4. 주문번호 개수 = unique QSM 주문번호 (enabled QSM)
    unique_orders = len({(q.get('주문번호') or '').strip() for q, _ in enabled_qsm
                         if (q.get('주문번호') or '').strip()})

    # 실제 출고 PCS = SUM of 予定数量 (QSM수량 × 매핑수량)
    total_picking_pcs = sum(int(r.get('予定数量') or 0) for r in outbound_rows)

    # outbound_rows 검증
    outbound_carts = len({r['注文番号'] for r in outbound_rows if r.get('注文番호')}) \
        if False else len({r['注文番号'] for r in outbound_rows if r.get('注文番号')})
    outbound_rows_count = len(outbound_rows)

    # 주문번호 개수 검증: 모든 enabled QSM 주문번호가 최소 1건 이상 출고 발생했는지
    # (실제로는 매핑 성공=enabled이면 무조건 생성되므로 거의 항상 True. sanity check)
    enabled_order_set = {(q.get('주문번호') or '').strip() for q, _ in enabled_qsm
                         if (q.get('주문번호') or '').strip()}
    # 출고 생성에 기여한 주문번호 (잠재적으로 매핑이 SKU=0행을 만들었을 때 drop될 수 있음)
    generated_orders = unique_orders  # = len(enabled_order_set) by construction
    check_orders_covered = generated_orders == len(enabled_order_set)

    return {
        'total_item_qty': total_item_qty,              # 총 상품 수량 (매핑 수량 합)
        'upload_row_count': upload_row_count,          # 주문 업로드 개수 (KSE row)
        'unique_carts': unique_carts,                  # 송장번호 개수
        'unique_orders': unique_orders,                # 주문번호 개수 (QSM)
        'total_picking_pcs': total_picking_pcs,        # 실제 출고 PCS (×QSM수량)
        'check_total_match_count': total_item_qty == upload_row_count,
        'check_carts_match': unique_carts == outbound_carts,
        'check_rows_match': upload_row_count == outbound_rows_count,
        'check_orders_covered': check_orders_covered,
    }


def build_outbound_xlsx(outbound_rows: List[Dict]) -> bytes:
    """
    원본 템플릿을 로드해 서식(컬럼 너비, 헤더 스타일, 폰트, 색상 등)을 그대로 보존한 채
    데이터 행만 교체하여 bytes 반환.
    """
    wb = openpyxl.load_workbook(OUTBOUND_TEMPLATE)
    ws = wb.active  # "Excel Sample"

    # 기존 샘플 데이터 행 삭제 (row 2 ~ max_row). 헤더(row1)는 유지.
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    # 헤더(row1)의 각 컬럼 셀 서식을 template_style로 기억
    # → 데이터 행의 기본 서식으로 사용할 수 있음. 그러나 원본에선
    # 데이터 행이 별도 서식(맑은 고딕 등)이었으므로 그걸 재현하기 위해
    # 원본 row2 스타일 템플릿을 미리 보관.
    # 이미 delete_rows로 지웠으므로, 원본을 다시 읽어서 row2 스타일을 가져온다.
    style_wb = openpyxl.load_workbook(OUTBOUND_TEMPLATE)
    style_ws = style_wb.active
    data_styles = []
    for c in range(1, style_ws.max_column + 1):
        src = style_ws.cell(2, c)
        data_styles.append({
            'font': copy.copy(src.font),
            'fill': copy.copy(src.fill),
            'alignment': copy.copy(src.alignment),
            'border': copy.copy(src.border),
            'number_format': src.number_format,
        })
    style_wb.close()

    # 데이터 행 추가
    for ridx, row in enumerate(outbound_rows, start=2):
        for c, (jp, _) in enumerate(OUTBOUND_HEADERS, 1):
            cell = ws.cell(ridx, c, row.get(jp, ''))
            s = data_styles[c - 1] if c - 1 < len(data_styles) else None
            if s:
                cell.font = s['font']
                cell.fill = s['fill']
                cell.alignment = s['alignment']
                cell.border = s['border']
                cell.number_format = s['number_format']

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def save_outbound_log(qsm_rows: List[Dict], outbound_rows: List[Dict],
                      mappings: Dict, source_file: str) -> int:
    """
    생성된 Outbound 데이터를 qoo10_outbound 테이블에 기록.
    (qoo10_cart_no, qoo10_order_no, sku_code) 기준 upsert.
    반환: 저장된 행 수
    """
    # 장바구니번호 → QSM 주문번호 (첫 매칭) + 원본 상품명/옵션/수량 매핑
    qsm_by_cart = {}
    for q in qsm_rows:
        cart = (q.get('장바구니번호') or '').strip()
        if cart and cart not in qsm_by_cart:
            qsm_by_cart[cart] = q

    conn = pg.connect()
    n = 0
    with conn.cursor() as cur:
        for row in outbound_rows:
            cart_no = str(row.get('注文番号', ''))
            sku_code = row.get('商品コード', '')
            qty = int(row.get('予定数量', 0) or 0)

            q_info = qsm_by_cart.get(cart_no, {})
            order_no = (q_info.get('주문번호') or '').strip()
            qsm_qty = int(q_info.get('수량', 1) or 1)
            qoo10_name = q_info.get('상품명', '')
            qoo10_option = q_info.get('옵션정보', '')

            # sku_name은 매핑에서 찾기
            sku_name = ''
            for m in mappings.values():
                for code, qn in zip(m['sku_codes'], m.get('item_codes', [])) \
                        if isinstance(m, dict) and 'item_codes' in m else []:
                    pass
            # 간단히: sku_catalog에서 역조회
            try:
                sku_list = load_kse_sku_catalog()
                sku_name = next((s['sku_name'] for s in sku_list if s['sku_code'] == sku_code), '')
            except Exception:
                sku_name = ''

            cur.execute("""
                INSERT INTO qoo10_outbound
                (qoo10_cart_no, qoo10_order_no, sku_code, sku_name, planned_qty,
                 recipient, recipient_phone, postal_code, address,
                 qoo10_product_name, qoo10_option, qoo10_qty, source_file, generated_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Seoul'))
                ON CONFLICT (qoo10_cart_no, qoo10_order_no, sku_code) DO UPDATE SET
                    sku_name = EXCLUDED.sku_name,
                    planned_qty = EXCLUDED.planned_qty,
                    recipient = EXCLUDED.recipient,
                    recipient_phone = EXCLUDED.recipient_phone,
                    postal_code = EXCLUDED.postal_code,
                    address = EXCLUDED.address,
                    qoo10_product_name = EXCLUDED.qoo10_product_name,
                    qoo10_option = EXCLUDED.qoo10_option,
                    qoo10_qty = EXCLUDED.qoo10_qty,
                    source_file = EXCLUDED.source_file,
                    generated_at = (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Seoul')
            """, (
                cart_no, order_no, sku_code, sku_name, qty,
                row.get('仕入先名/受取人名', ''),
                row.get('電話番号', ''),
                row.get('郵便番号コード', ''),
                row.get('基本住所', ''),
                qoo10_name, qoo10_option, qsm_qty, source_file,
            ))
            n += 1
    conn.commit()
    conn.close()
    return n


def update_outbound_waybills(waybill_map: Dict[str, str]) -> int:
    """장바구니번호 → 송장번호 매핑으로 qoo10_outbound.waybill 갱신.
    반환: 갱신된 행 수
    """
    if not waybill_map:
        return 0
    conn = pg.connect()
    total = 0
    with conn.cursor() as cur:
        for cart, waybill in waybill_map.items():
            cur.execute("""
                UPDATE qoo10_outbound
                SET waybill = %s, waybill_updated_at = (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Seoul')
                WHERE qoo10_cart_no = %s
            """, (waybill, cart))
            total += cur.rowcount
    conn.commit()
    conn.close()
    return total


def count_disabled_in_brief(brief_rows: List[Dict], mappings: Dict) -> int:
    """
    brief.csv 행들 중 (상품명+옵션) 매핑이 '취급안함(enabled=False)'에 해당하는 행 수.
    Tab ② 단독에서 Tab ①과 동일한 검수가 가능하도록 함.
    """
    n = 0
    for r in brief_rows:
        name = (r.get('상품명') or '').strip()
        option = (r.get('옵션정보') or '').strip()
        m = mappings.get((name, option))
        if m and not m['enabled']:
            n += 1
    return n


def save_pending_brief(content: bytes, file_name: str, cart_count: int,
                        disabled_count: int = 0) -> int:
    """brief.csv 바이트를 DB에 임시저장. 이미 같은 파일명이 있으면 덮어쓰기.
    disabled_count: Tab ① 취급안함(매핑 비활성) 분류 건수 — Tab ②에서 기대 미취급 수로 활용.
    """
    conn = pg.connect()
    with conn.cursor() as cur:
        cur.execute("""
            SELECT id FROM qoo10_pending_brief
            WHERE file_name = %s AND consumed_at IS NULL
            ORDER BY created_at DESC LIMIT 1
        """, (file_name,))
        existing = cur.fetchone()
        if existing:
            cur.execute("""
                UPDATE qoo10_pending_brief
                SET content = %s, cart_count = %s, disabled_count = %s,
                    created_at = (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Seoul')
                WHERE id = %s
            """, (content, cart_count, disabled_count, existing[0]))
            rid = existing[0]
        else:
            cur.execute("""
                INSERT INTO qoo10_pending_brief (file_name, content, cart_count, disabled_count)
                VALUES (%s, %s, %s, %s) RETURNING id
            """, (file_name, content, cart_count, disabled_count))
            rid = cur.fetchone()[0]
    conn.commit()
    conn.close()
    return rid


def list_pending_briefs(include_consumed: bool = False, limit: int = 20) -> List[Dict]:
    """임시저장된 brief 목록"""
    conn = pg.connect(autocommit=True)
    with conn.cursor() as cur:
        where = "" if include_consumed else "WHERE consumed_at IS NULL"
        cur.execute(f"""
            SELECT id, created_at, file_name, cart_count, disabled_count, consumed_at
            FROM qoo10_pending_brief {where}
            ORDER BY created_at DESC LIMIT %s
        """, (limit,))
        rows = cur.fetchall()
    conn.close()
    return [
        {'id': r[0], 'created_at': r[1], 'file_name': r[2],
         'cart_count': r[3], 'disabled_count': r[4] or 0, 'consumed_at': r[5]}
        for r in rows
    ]


def load_pending_brief(brief_id: int) -> Tuple[bytes, str]:
    """특정 임시저장 brief 로드"""
    conn = pg.connect(autocommit=True)
    with conn.cursor() as cur:
        cur.execute(
            "SELECT content, file_name FROM qoo10_pending_brief WHERE id = %s",
            (brief_id,),
        )
        row = cur.fetchone()
    conn.close()
    if not row:
        raise RuntimeError(f"임시저장 brief id={brief_id} 없음")
    return bytes(row[0]), row[1]


def mark_brief_consumed(brief_id: int):
    """임시저장 brief를 consumed로 표시 (송장 업로드 완료 후)"""
    conn = pg.connect()
    with conn.cursor() as cur:
        cur.execute("""
            UPDATE qoo10_pending_brief
            SET consumed_at = (CURRENT_TIMESTAMP AT TIME ZONE 'Asia/Seoul')
            WHERE id = %s
        """, (brief_id,))
    conn.commit()
    conn.close()


def delete_pending_brief(brief_id: int):
    """임시저장 brief 삭제 (사용자가 취소 선택 시)"""
    conn = pg.connect()
    with conn.cursor() as cur:
        cur.execute("DELETE FROM qoo10_pending_brief WHERE id = %s", (brief_id,))
    conn.commit()
    conn.close()


def parse_kse_oms_waybill(xlsx_bytes: bytes) -> Dict[str, str]:
    """
    KSE OMS "주문(출고&입고) 내역" xlsx 파일 → {주문번호: 송장번호} 매핑.
    주문번호(col8) = QSM 장바구니번호(=Outbound 注文番号)
    운송장번호(col42) = Slack/QSM에 넘길 송장번호
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active
    # 헤더 위치 검증 (보통 row1)
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    try:
        order_col = hdr.index('주문 번호') + 1
        waybill_col = hdr.index('운송장 번호') + 1
        cancel_col = hdr.index('주문 취소 여부') + 1 if '주문 취소 여부' in hdr else None
    except ValueError:
        wb.close()
        raise RuntimeError("'주문 번호' 또는 '운송장 번호' 컬럼을 찾을 수 없습니다")

    mapping = {}
    for r in range(2, ws.max_row + 1):
        order_no = ws.cell(r, order_col).value
        waybill = ws.cell(r, waybill_col).value
        cancelled = ws.cell(r, cancel_col).value if cancel_col else None
        if order_no and waybill and str(cancelled).strip() != '네':
            mapping[str(order_no).strip()] = str(waybill).strip()
    wb.close()
    return mapping


def build_qsm_waybill_csv(brief_content: bytes, waybill_map: Dict[str, str]) -> Tuple[bytes, List[str]]:
    """
    brief.csv bytes + 장바구니번호→송장번호 매핑 → QSM 업로드용 CSV bytes.
    **원본 서식(BOM, 라인 엔딩, 미변경 행의 바이트)을 최대한 그대로 보존**.
    waybill 채울 행만 재직렬화, 나머지 행은 원본 라인 그대로.
    waybill_map: {장바구니번호: 송장번호}
    """
    has_bom = brief_content.startswith(b'\xef\xbb\xbf')
    text = brief_content.decode('utf-8-sig')

    # 원본 라인 엔딩 감지
    if '\r\n' in text:
        sep = '\r\n'
    elif '\r' in text:
        sep = '\r'
    else:
        sep = '\n'

    lines = text.split(sep)
    # 마지막 빈 줄 보존
    trailing_empty = bool(lines) and lines[-1] == ''
    if trailing_empty:
        lines = lines[:-1]

    if not lines:
        return brief_content, []

    # 헤더에서 컬럼 위치 찾기
    header_line = lines[0]
    headers = next(csv.reader(io.StringIO(header_line)))
    try:
        cart_idx = headers.index('장바구니번호')
        waybill_idx = headers.index('송장번호')
    except ValueError:
        return brief_content, []

    out_lines = [header_line]
    missing = []
    for line in lines[1:]:
        if not line:
            out_lines.append(line)
            continue
        fields = next(csv.reader(io.StringIO(line)))
        if len(fields) <= max(cart_idx, waybill_idx):
            out_lines.append(line)
            continue
        cart_no = (fields[cart_idx] or '').strip()
        wb = waybill_map.get(cart_no)
        if wb:
            # 해당 행만 송장번호 채워 재직렬화
            fields[waybill_idx] = wb
            buf = io.StringIO()
            csv.writer(buf, lineterminator='').writerow(fields)
            out_lines.append(buf.getvalue())
        else:
            # 미매칭은 원본 라인 그대로 (바이트 손실 0)
            out_lines.append(line)
            missing.append(cart_no)

    result = sep.join(out_lines)
    if trailing_empty:
        result += sep

    out_bytes = result.encode('utf-8')
    if has_bom:
        out_bytes = b'\xef\xbb\xbf' + out_bytes
    return out_bytes, missing
